import openpyxl
import json


def read_excel(file, sheet):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheet]
    return [(sh.cell(row=i, column=1).value, sh.cell(row=i, column=2).value)
            for i in range(2, sh.max_row + 1)]


def read_json(file):
    with open(file) as f:
        return json.load(f)
